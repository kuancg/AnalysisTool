"""
Tracker Benchmark Web Application
目标跟踪评测Web应用
"""
import os
import json
import uuid
import zipfile
import shutil
import tempfile
import traceback
import pickle
import numpy as np
from flask import Flask, request, jsonify, render_template, send_file, make_response

from evaluation import (
    load_txt_results, evaluate_multiple_trackers,
    compute_sequence_iou, compute_sequence_center_error
)
from visualization import (
    plot_success_curve, plot_precision_curve, plot_norm_precision_curve,
    plot_attribute_success, plot_attribute_precision, plot_radar_chart,
    plot_sequence_iou_curve, plot_sequence_center_error_curve,
    generate_all_plots
)

# Excel支持
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 图像处理支持
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB

# 数据持久化目录
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saved_data')
os.makedirs(DATA_DIR, exist_ok=True)

# 数据存储
DATA_STORE = {
    'sequence_lists': {},      # {dataset_name: [seq1, seq2, ...]} 存储上传的序列列表
    'ground_truths': {},      # {dataset_name: {seq_name: gt_bb}}
    'tracker_results': {},    # {dataset_name: {tracker_name: {seq_name: pred_bb}}}
    'tracker_times': {},      # {dataset_name: {tracker_name: {seq_name: times}}}  # 新增：存储时间数据
    'attributes': {},         # {dataset_name: {seq_name: att_vector}}
    'attribute_names': {},    # {dataset_name: [attr1, attr2, ...]}
    'evaluation_results': {}, # {session_id: results}
}


def save_data_to_disk():
    """保存数据到磁盘"""
    try:
        # 保存每种数据类型
        data_to_save = {
            'sequence_lists': DATA_STORE['sequence_lists'],
            'ground_truths': {ds: {seq: arr.tolist() for seq, arr in seqs.items()} 
                            for ds, seqs in DATA_STORE['ground_truths'].items()},
            'tracker_results': {ds: {trk: {seq: arr.tolist() for seq, arr in seqs.items()} 
                                    for trk, seqs in trks.items()} 
                               for ds, trks in DATA_STORE['tracker_results'].items()},
            'tracker_times': {ds: {trk: {seq: arr.tolist() if isinstance(arr, np.ndarray) else arr 
                                        for seq, arr in seqs.items()} 
                                  for trk, seqs in trks.items()} 
                             for ds, trks in DATA_STORE['tracker_times'].items()},
            'attributes': {ds: {seq: arr.tolist() if isinstance(arr, np.ndarray) else arr 
                               for seq, arr in seqs.items()} 
                          for ds, seqs in DATA_STORE['attributes'].items()},
            'attribute_names': DATA_STORE['attribute_names'],
        }
        
        save_path = os.path.join(DATA_DIR, 'benchmark_data.pkl')
        with open(save_path, 'wb') as f:
            pickle.dump(data_to_save, f)
        
        print(f"[INFO] 数据已保存到 {save_path}")
        return True
    except Exception as e:
        print(f"[ERROR] 保存数据失败: {e}")
        return False


def load_data_from_disk():
    """从磁盘加载数据"""
    try:
        save_path = os.path.join(DATA_DIR, 'benchmark_data.pkl')
        if not os.path.exists(save_path):
            print("[INFO] 没有找到已保存的数据文件")
            return False
        
        with open(save_path, 'rb') as f:
            data = pickle.load(f)
        
        # 恢复数据，将list转回numpy数组
        DATA_STORE['sequence_lists'] = data.get('sequence_lists', {})
        
        DATA_STORE['ground_truths'] = {
            ds: {seq: np.array(arr) for seq, arr in seqs.items()} 
            for ds, seqs in data.get('ground_truths', {}).items()
        }
        
        DATA_STORE['tracker_results'] = {
            ds: {trk: {seq: np.array(arr) for seq, arr in seqs.items()} 
                 for trk, seqs in trks.items()} 
            for ds, trks in data.get('tracker_results', {}).items()
        }
        
        DATA_STORE['tracker_times'] = {
            ds: {trk: {seq: np.array(arr) if isinstance(arr, list) else arr 
                      for seq, arr in seqs.items()} 
                 for trk, seqs in trks.items()} 
            for ds, trks in data.get('tracker_times', {}).items()
        }
        
        DATA_STORE['attributes'] = {
            ds: {seq: np.array(arr) if isinstance(arr, list) else arr 
                 for seq, arr in seqs.items()} 
            for ds, seqs in data.get('attributes', {}).items()
        }
        
        DATA_STORE['attribute_names'] = data.get('attribute_names', {})
        
        # 统计信息
        gt_count = sum(len(seqs) for seqs in DATA_STORE['ground_truths'].values())
        trk_count = sum(len(trks) for trks in DATA_STORE['tracker_results'].values())
        ds_count = len(DATA_STORE['ground_truths'])
        
        print(f"[INFO] 已从 {save_path} 加载数据:")
        print(f"       - {ds_count} 个数据集")
        print(f"       - {gt_count} 个GT序列")
        print(f"       - {trk_count} 个跟踪器")
        
        return True
    except Exception as e:
        print(f"[ERROR] 加载数据失败: {e}")
        traceback.print_exc()
        return False


# 启动时自动加载数据
load_data_from_disk()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/status', methods=['GET'])
def get_status():
    """获取当前数据状态"""
    datasets = list(DATA_STORE['ground_truths'].keys())
    
    trackers_by_dataset = {}
    for ds_name, trackers in DATA_STORE['tracker_results'].items():
        trackers_by_dataset[ds_name] = list(trackers.keys())
    
    attributes_by_dataset = {}
    for ds_name, attrs in DATA_STORE['attribute_names'].items():
        attributes_by_dataset[ds_name] = attrs
    
    gt_counts = {}
    sequences_by_dataset = {}
    for ds_name, seqs in DATA_STORE['ground_truths'].items():
        gt_counts[ds_name] = len(seqs)
        sequences_by_dataset[ds_name] = sorted(list(seqs.keys()))
    
    # 序列列表信息
    sequence_lists_info = {}
    for ds_name, seq_list in DATA_STORE['sequence_lists'].items():
        sequence_lists_info[ds_name] = {
            'count': len(seq_list),
            'sequences': seq_list
        }
    
    # 检查是否有已保存的数据文件
    save_path = os.path.join(DATA_DIR, 'benchmark_data.pkl')
    has_saved_data = os.path.exists(save_path)
    saved_data_info = None
    if has_saved_data:
        try:
            import time
            mtime = os.path.getmtime(save_path)
            size_mb = os.path.getsize(save_path) / (1024 * 1024)
            saved_data_info = {
                'last_modified': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime)),
                'size_mb': round(size_mb, 2)
            }
        except:
            pass
    
    return jsonify({
        'datasets': datasets,
        'trackers_by_dataset': trackers_by_dataset,
        'attributes_by_dataset': attributes_by_dataset,
        'gt_counts': gt_counts,
        'sequences_by_dataset': sequences_by_dataset,
        'sequence_lists': sequence_lists_info,
        'has_saved_data': has_saved_data,
        'saved_data_info': saved_data_info
    })


@app.route('/api/clear', methods=['POST'])
def clear_data():
    """清空所有数据"""
    data = request.get_json() or {}
    delete_saved = data.get('delete_saved', False)
    
    DATA_STORE['sequence_lists'] = {}
    DATA_STORE['ground_truths'] = {}
    DATA_STORE['tracker_results'] = {}
    DATA_STORE['tracker_times'] = {}
    DATA_STORE['attributes'] = {}
    DATA_STORE['attribute_names'] = {}
    DATA_STORE['evaluation_results'] = {}
    
    message = '内存中的数据已清空'
    
    if delete_saved:
        save_path = os.path.join(DATA_DIR, 'benchmark_data.pkl')
        if os.path.exists(save_path):
            os.remove(save_path)
            message += '，已保存的数据文件也已删除'
        else:
            message += '（没有找到已保存的数据文件）'
    else:
        message += '（已保存的数据文件保留，下次启动时会自动加载）'
    
    return jsonify({'success': True, 'message': message})


@app.route('/api/upload/sequence_list', methods=['POST'])
def upload_sequence_list():
    """上传视频序列列表信息"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    
    file = request.files['file']
    mode = request.form.get('mode', 'single')
    dataset_name = request.form.get('dataset_name', '')
    
    temp_dir = tempfile.mkdtemp()
    try:
        if file.filename.endswith('.zip'):
            # 批量模式：ZIP包含多个txt文件，每个以数据集名命名
            zip_path = os.path.join(temp_dir, 'upload.zip')
            file.save(zip_path)
            
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(temp_dir)
            
            extract_dir = temp_dir
            items = os.listdir(extract_dir)
            items = [i for i in items if not i.startswith('.') and not i.startswith('__') and i != 'upload.zip']
            
            # 如果只有一个文件夹，进入该文件夹
            if len(items) == 1 and os.path.isdir(os.path.join(extract_dir, items[0])):
                extract_dir = os.path.join(extract_dir, items[0])
                items = os.listdir(extract_dir)
                items = [i for i in items if not i.startswith('.') and not i.startswith('__')]
            
            loaded_info = []
            for item in items:
                item_path = os.path.join(extract_dir, item)
                if item.endswith('.txt') and os.path.isfile(item_path):
                    ds_name = item[:-4]  # 去掉.txt后缀作为数据集名
                    try:
                        with open(item_path, 'r', encoding='utf-8') as f:
                            seq_list = [line.strip() for line in f if line.strip()]
                        if seq_list:
                            DATA_STORE['sequence_lists'][ds_name] = seq_list
                            loaded_info.append(f'{ds_name}: {len(seq_list)}个序列')
                    except Exception as e:
                        pass
            
            if not loaded_info:
                return jsonify({'error': '未找到有效的txt文件，请检查压缩包结构'}), 400
            
            # 保存数据到磁盘
            save_data_to_disk()
            
            return jsonify({
                'success': True,
                'message': '视频序列列表上传成功（批量模式，已自动保存）',
                'details': loaded_info
            })
        
        elif file.filename.endswith('.txt'):
            # 单数据集模式：单个txt文件
            if not dataset_name:
                return jsonify({'error': '单数据集模式需要输入数据集名称'}), 400
            
            txt_path = os.path.join(temp_dir, 'sequences.txt')
            file.save(txt_path)
            
            with open(txt_path, 'r', encoding='utf-8') as f:
                seq_list = [line.strip() for line in f if line.strip()]
            
            if not seq_list:
                return jsonify({'error': 'txt文件为空或格式不正确'}), 400
            
            DATA_STORE['sequence_lists'][dataset_name] = seq_list
            
            # 保存数据到磁盘
            save_data_to_disk()
            
            return jsonify({
                'success': True,
                'message': '视频序列列表上传成功（已自动保存）',
                'details': [f'{dataset_name}: {len(seq_list)}个序列']
            })
        
        else:
            return jsonify({'error': '请上传ZIP或TXT格式文件'}), 400
    
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}', 'traceback': traceback.format_exc()}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def verify_sequences_against_list(dataset_name: str, actual_sequences: list, data_type: str) -> dict:
    """验证上传的序列是否与预定义的序列列表一致
    
    Args:
        dataset_name: 数据集名称
        actual_sequences: 实际上传的序列名列表
        data_type: 数据类型描述（如 'Ground Truth', '跟踪结果', '属性'）
    
    Returns:
        dict: {
            'verified': bool,
            'missing': list,  # 在序列列表中但未上传的序列
            'extra': list,    # 上传了但不在序列列表中的序列
            'message': str
        }
    """
    if dataset_name not in DATA_STORE['sequence_lists']:
        return {
            'verified': False,
            'has_list': False,
            'message': f'数据集 {dataset_name} 没有上传序列列表，无法验证'
        }
    
    expected_sequences = set(DATA_STORE['sequence_lists'][dataset_name])
    actual_set = set(actual_sequences)
    
    missing = list(expected_sequences - actual_set)
    extra = list(actual_set - expected_sequences)
    
    result = {
        'verified': len(missing) == 0 and len(extra) == 0,
        'has_list': True,
        'expected_count': len(expected_sequences),
        'actual_count': len(actual_set),
        'missing': sorted(missing),
        'extra': sorted(extra),
        'missing_count': len(missing),
        'extra_count': len(extra)
    }
    
    if result['verified']:
        result['message'] = f'{data_type}验证通过：{len(actual_set)}个序列与序列列表完全匹配'
    else:
        msgs = []
        if missing:
            msgs.append(f'缺少 {len(missing)} 个序列')
        if extra:
            msgs.append(f'多余 {len(extra)} 个序列')
        result['message'] = f'{data_type}验证失败：' + '，'.join(msgs)
    
    return result


@app.route('/api/upload/groundtruth', methods=['POST'])
def upload_groundtruth():
    """上传Ground Truth"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.zip'):
        return jsonify({'error': '请上传ZIP格式文件'}), 400
    
    mode = request.form.get('mode', 'single')
    dataset_name = request.form.get('dataset_name', '')
    
    temp_dir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(temp_dir, 'upload.zip')
        file.save(zip_path)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
        
        extract_dir = temp_dir
        items = os.listdir(extract_dir)
        items = [i for i in items if not i.startswith('.') and not i.startswith('__') and i != 'upload.zip']
        
        if len(items) == 1 and os.path.isdir(os.path.join(extract_dir, items[0])):
            extract_dir = os.path.join(extract_dir, items[0])
            items = os.listdir(extract_dir)
            items = [i for i in items if not i.startswith('.') and not i.startswith('__')]
        
        loaded_info = []
        
        if mode == 'batch':
            # 批量模式：每个子文件夹是一个数据集
            for item in items:
                item_path = os.path.join(extract_dir, item)
                if os.path.isdir(item_path):
                    ds_name = item
                    if ds_name not in DATA_STORE['ground_truths']:
                        DATA_STORE['ground_truths'][ds_name] = {}
                    
                    count = 0
                    for txt_file in os.listdir(item_path):
                        if txt_file.endswith('.txt'):
                            seq_name = txt_file[:-4]
                            try:
                                gt_data = load_txt_results(os.path.join(item_path, txt_file))
                                DATA_STORE['ground_truths'][ds_name][seq_name] = gt_data
                                count += 1
                            except Exception as e:
                                pass
                    
                    if count > 0:
                        loaded_info.append(f'{ds_name}: {count}个序列')
        else:
            # 单数据集模式
            if not dataset_name:
                return jsonify({'error': '单数据集模式需要输入数据集名称'}), 400
            
            if dataset_name not in DATA_STORE['ground_truths']:
                DATA_STORE['ground_truths'][dataset_name] = {}
            
            count = 0
            for root, dirs, files in os.walk(extract_dir):
                for txt_file in files:
                    if txt_file.endswith('.txt'):
                        seq_name = txt_file[:-4]
                        try:
                            gt_data = load_txt_results(os.path.join(root, txt_file))
                            DATA_STORE['ground_truths'][dataset_name][seq_name] = gt_data
                            count += 1
                        except Exception as e:
                            pass
            
            loaded_info.append(f'{dataset_name}: {count}个序列')
        
        if not loaded_info:
            return jsonify({'error': '未找到有效的txt文件，请检查压缩包结构'}), 400
        
        # 验证序列与序列列表的一致性
        verification_results = {}
        all_verified = True
        
        if mode == 'batch':
            for ds_name in DATA_STORE['ground_truths'].keys():
                if ds_name in DATA_STORE['sequence_lists']:
                    actual_seqs = list(DATA_STORE['ground_truths'][ds_name].keys())
                    verify_result = verify_sequences_against_list(ds_name, actual_seqs, 'Ground Truth')
                    verification_results[ds_name] = verify_result
                    if not verify_result.get('verified', True):
                        all_verified = False
        else:
            if dataset_name in DATA_STORE['sequence_lists']:
                actual_seqs = list(DATA_STORE['ground_truths'][dataset_name].keys())
                verify_result = verify_sequences_against_list(dataset_name, actual_seqs, 'Ground Truth')
                verification_results[dataset_name] = verify_result
                if not verify_result.get('verified', True):
                    all_verified = False
        
        # 保存数据到磁盘
        save_data_to_disk()
        
        response = {
            'success': True,
            'message': f'Ground Truth上传成功（已自动保存）',
            'details': loaded_info
        }
        
        if verification_results:
            response['verification'] = verification_results
            response['all_verified'] = all_verified
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}', 'traceback': traceback.format_exc()}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.route('/api/upload/tracking_results', methods=['POST'])
def upload_tracking_results():
    """上传跟踪结果"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.zip'):
        return jsonify({'error': '请上传ZIP格式文件'}), 400
    
    mode = request.form.get('mode', 'single')
    tracker_name = request.form.get('tracker_name', '')
    dataset_name = request.form.get('dataset_name', '')
    
    temp_dir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(temp_dir, 'upload.zip')
        file.save(zip_path)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
        
        extract_dir = temp_dir
        items = os.listdir(extract_dir)
        items = [i for i in items if not i.startswith('.') and not i.startswith('__') and i != 'upload.zip']
        
        if len(items) == 1 and os.path.isdir(os.path.join(extract_dir, items[0])):
            extract_dir = os.path.join(extract_dir, items[0])
            items = os.listdir(extract_dir)
            items = [i for i in items if not i.startswith('.') and not i.startswith('__')]
        
        loaded_info = []
        
        def load_time_file(filepath):
            """加载时间文件"""
            try:
                times = []
                with open(filepath, 'r') as f:
                    for line in f:
                        line = line.strip()
                        if line:
                            times.append(float(line))
                return np.array(times) if times else None
            except:
                return None
        
        if mode == 'batch':
            # 批量模式：tracker_name/dataset_name/seq.txt（多算法多数据集）
            for trk_item in items:
                trk_path = os.path.join(extract_dir, trk_item)
                if not os.path.isdir(trk_path):
                    continue
                
                trk_name = trk_item
                
                for ds_item in os.listdir(trk_path):
                    ds_path = os.path.join(trk_path, ds_item)
                    if not os.path.isdir(ds_path):
                        continue
                    
                    ds_name = ds_item
                    
                    if ds_name not in DATA_STORE['tracker_results']:
                        DATA_STORE['tracker_results'][ds_name] = {}
                    if trk_name not in DATA_STORE['tracker_results'][ds_name]:
                        DATA_STORE['tracker_results'][ds_name][trk_name] = {}
                    
                    if ds_name not in DATA_STORE['tracker_times']:
                        DATA_STORE['tracker_times'][ds_name] = {}
                    if trk_name not in DATA_STORE['tracker_times'][ds_name]:
                        DATA_STORE['tracker_times'][ds_name][trk_name] = {}
                    
                    count = 0
                    time_count = 0
                    for txt_file in os.listdir(ds_path):
                        if txt_file.endswith('.txt'):
                            # 检查是否是时间文件
                            if txt_file.lower() == 'time.txt' or txt_file.endswith('_time.txt'):
                                seq_name = txt_file[:-9] if txt_file.endswith('_time.txt') else None
                                time_data = load_time_file(os.path.join(ds_path, txt_file))
                                if time_data is not None and seq_name:
                                    DATA_STORE['tracker_times'][ds_name][trk_name][seq_name] = time_data
                                    time_count += 1
                            else:
                                seq_name = txt_file[:-4]
                                try:
                                    pred_data = load_txt_results(os.path.join(ds_path, txt_file))
                                    DATA_STORE['tracker_results'][ds_name][trk_name][seq_name] = pred_data
                                    count += 1
                                    
                                    # 尝试加载对应的时间文件
                                    time_file = os.path.join(ds_path, seq_name + '_time.txt')
                                    if os.path.exists(time_file):
                                        time_data = load_time_file(time_file)
                                        if time_data is not None:
                                            DATA_STORE['tracker_times'][ds_name][trk_name][seq_name] = time_data
                                            time_count += 1
                                except:
                                    pass
                    
                    if count > 0:
                        info = f'{trk_name}/{ds_name}: {count}个序列'
                        if time_count > 0:
                            info += f', {time_count}个时间文件'
                        loaded_info.append(info)
        
        elif mode == 'single_tracker':
            # 单算法多数据集模式：tracker_name/dataset_name/seq.txt
            # 需要提供算法名称，压缩包内直接是数据集文件夹
            if not tracker_name:
                return jsonify({'error': '单算法多数据集模式需要输入算法名称'}), 400
            
            trk_name = tracker_name
            
            for ds_item in items:
                ds_path = os.path.join(extract_dir, ds_item)
                if not os.path.isdir(ds_path):
                    continue
                
                ds_name = ds_item
                
                if ds_name not in DATA_STORE['tracker_results']:
                    DATA_STORE['tracker_results'][ds_name] = {}
                if trk_name not in DATA_STORE['tracker_results'][ds_name]:
                    DATA_STORE['tracker_results'][ds_name][trk_name] = {}
                
                if ds_name not in DATA_STORE['tracker_times']:
                    DATA_STORE['tracker_times'][ds_name] = {}
                if trk_name not in DATA_STORE['tracker_times'][ds_name]:
                    DATA_STORE['tracker_times'][ds_name][trk_name] = {}
                
                count = 0
                time_count = 0
                for txt_file in os.listdir(ds_path):
                    if txt_file.endswith('.txt'):
                        # 检查是否是时间文件
                        if txt_file.lower() == 'time.txt' or txt_file.endswith('_time.txt'):
                            seq_name = txt_file[:-9] if txt_file.endswith('_time.txt') else None
                            time_data = load_time_file(os.path.join(ds_path, txt_file))
                            if time_data is not None and seq_name:
                                DATA_STORE['tracker_times'][ds_name][trk_name][seq_name] = time_data
                                time_count += 1
                        else:
                            seq_name = txt_file[:-4]
                            try:
                                pred_data = load_txt_results(os.path.join(ds_path, txt_file))
                                DATA_STORE['tracker_results'][ds_name][trk_name][seq_name] = pred_data
                                count += 1
                                
                                # 尝试加载对应的时间文件
                                time_file = os.path.join(ds_path, seq_name + '_time.txt')
                                if os.path.exists(time_file):
                                    time_data = load_time_file(time_file)
                                    if time_data is not None:
                                        DATA_STORE['tracker_times'][ds_name][trk_name][seq_name] = time_data
                                        time_count += 1
                            except:
                                pass
                
                if count > 0:
                    info = f'{trk_name}/{ds_name}: {count}个序列'
                    if time_count > 0:
                        info += f', {time_count}个时间文件'
                    loaded_info.append(info)
        
        else:
            # 单算法单数据集模式
            if not tracker_name:
                return jsonify({'error': '单上传模式需要输入算法名称'}), 400
            if not dataset_name:
                return jsonify({'error': '单上传模式需要输入数据集名称'}), 400
            
            if dataset_name not in DATA_STORE['tracker_results']:
                DATA_STORE['tracker_results'][dataset_name] = {}
            if tracker_name not in DATA_STORE['tracker_results'][dataset_name]:
                DATA_STORE['tracker_results'][dataset_name][tracker_name] = {}
            
            if dataset_name not in DATA_STORE['tracker_times']:
                DATA_STORE['tracker_times'][dataset_name] = {}
            if tracker_name not in DATA_STORE['tracker_times'][dataset_name]:
                DATA_STORE['tracker_times'][dataset_name][tracker_name] = {}
            
            count = 0
            time_count = 0
            for root, dirs, files in os.walk(extract_dir):
                for txt_file in files:
                    if txt_file.endswith('.txt'):
                        if txt_file.lower() == 'time.txt' or txt_file.endswith('_time.txt'):
                            continue  # 跳过时间文件，后面单独处理
                        
                        seq_name = txt_file[:-4]
                        try:
                            pred_data = load_txt_results(os.path.join(root, txt_file))
                            DATA_STORE['tracker_results'][dataset_name][tracker_name][seq_name] = pred_data
                            count += 1
                            
                            # 尝试加载对应的时间文件
                            time_file = os.path.join(root, seq_name + '_time.txt')
                            if os.path.exists(time_file):
                                time_data = load_time_file(time_file)
                                if time_data is not None:
                                    DATA_STORE['tracker_times'][dataset_name][tracker_name][seq_name] = time_data
                                    time_count += 1
                        except:
                            pass
            
            info = f'{tracker_name}/{dataset_name}: {count}个序列'
            if time_count > 0:
                info += f', {time_count}个时间文件'
            loaded_info.append(info)
        
        if not loaded_info:
            return jsonify({'error': '未找到有效的txt文件，请检查压缩包结构'}), 400
        
        # 验证序列与序列列表的一致性
        verification_results = {}
        all_verified = True
        
        if mode == 'batch':
            # 批量模式：遍历所有数据集的所有跟踪器
            for ds_name, trackers in DATA_STORE['tracker_results'].items():
                if ds_name in DATA_STORE['sequence_lists']:
                    for trk_name, seqs in trackers.items():
                        actual_seqs = list(seqs.keys())
                        verify_result = verify_sequences_against_list(
                            ds_name, actual_seqs, f'跟踪结果({trk_name})')
                        key = f'{trk_name}/{ds_name}'
                        verification_results[key] = verify_result
                        if not verify_result.get('verified', True):
                            all_verified = False
        elif mode == 'single_tracker':
            # 单算法多数据集模式：验证该算法在各数据集上的结果
            if tracker_name in [trk for ds in DATA_STORE['tracker_results'].values() for trk in ds.keys()]:
                for ds_name, trackers in DATA_STORE['tracker_results'].items():
                    if tracker_name in trackers and ds_name in DATA_STORE['sequence_lists']:
                        actual_seqs = list(trackers[tracker_name].keys())
                        verify_result = verify_sequences_against_list(
                            ds_name, actual_seqs, f'跟踪结果({tracker_name})')
                        key = f'{tracker_name}/{ds_name}'
                        verification_results[key] = verify_result
                        if not verify_result.get('verified', True):
                            all_verified = False
        else:
            # 单算法单数据集模式
            if dataset_name in DATA_STORE['sequence_lists']:
                actual_seqs = list(DATA_STORE['tracker_results'][dataset_name][tracker_name].keys())
                verify_result = verify_sequences_against_list(
                    dataset_name, actual_seqs, f'跟踪结果({tracker_name})')
                key = f'{tracker_name}/{dataset_name}'
                verification_results[key] = verify_result
                if not verify_result.get('verified', True):
                    all_verified = False
        
        # 保存数据到磁盘
        save_data_to_disk()
        
        response = {
            'success': True,
            'message': '跟踪结果上传成功（已自动保存）',
            'details': loaded_info
        }
        
        if verification_results:
            response['verification'] = verification_results
            response['all_verified'] = all_verified
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}', 'traceback': traceback.format_exc()}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.route('/api/upload/attributes', methods=['POST'])
def upload_attributes():
    """上传属性配置"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.zip'):
        return jsonify({'error': '请上传ZIP格式文件'}), 400
    
    mode = request.form.get('mode', 'single')
    dataset_name = request.form.get('dataset_name', '')
    attribute_names_str = request.form.get('attribute_names', '')
    
    temp_dir = tempfile.mkdtemp()
    try:
        zip_path = os.path.join(temp_dir, 'upload.zip')
        file.save(zip_path)
        
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(temp_dir)
        
        extract_dir = temp_dir
        items = os.listdir(extract_dir)
        items = [i for i in items if not i.startswith('.') and not i.startswith('__') and i != 'upload.zip']
        
        if len(items) == 1 and os.path.isdir(os.path.join(extract_dir, items[0])):
            extract_dir = os.path.join(extract_dir, items[0])
            items = os.listdir(extract_dir)
            items = [i for i in items if not i.startswith('.') and not i.startswith('__')]
        
        loaded_info = []
        
        if mode == 'batch':
            # 批量模式：每个子文件夹是一个数据集
            for item in items:
                item_path = os.path.join(extract_dir, item)
                if not os.path.isdir(item_path):
                    continue
                
                ds_name = item
                
                # 读取att_info.txt获取属性名称
                att_info_path = os.path.join(item_path, 'att_info.txt')
                if os.path.exists(att_info_path):
                    with open(att_info_path, 'r', encoding='utf-8') as f:
                        content = f.read().strip()
                        attr_names = [a.strip() for a in content.split(',') if a.strip()]
                else:
                    continue
                
                if not attr_names:
                    continue
                
                DATA_STORE['attribute_names'][ds_name] = attr_names
                
                if ds_name not in DATA_STORE['attributes']:
                    DATA_STORE['attributes'][ds_name] = {}
                
                count = 0
                for txt_file in os.listdir(item_path):
                    if txt_file.endswith('.txt') and txt_file != 'att_info.txt':
                        seq_name = txt_file[:-4]
                        try:
                            with open(os.path.join(item_path, txt_file), 'r') as f:
                                content = f.read().strip()
                                att_vec = [int(x.strip()) for x in content.split(',') if x.strip()]
                                DATA_STORE['attributes'][ds_name][seq_name] = np.array(att_vec)
                                count += 1
                        except:
                            pass
                
                if count > 0:
                    loaded_info.append(f'{ds_name}: {count}个序列, {len(attr_names)}个属性 ({", ".join(attr_names)})')
        else:
            # 单数据集模式
            if not dataset_name:
                return jsonify({'error': '单数据集模式需要输入数据集名称'}), 400
            if not attribute_names_str:
                return jsonify({'error': '单数据集模式需要输入属性名称（逗号分隔）'}), 400
            
            attr_names = [a.strip() for a in attribute_names_str.split(',') if a.strip()]
            if not attr_names:
                return jsonify({'error': '属性名称格式错误'}), 400
            
            DATA_STORE['attribute_names'][dataset_name] = attr_names
            
            if dataset_name not in DATA_STORE['attributes']:
                DATA_STORE['attributes'][dataset_name] = {}
            
            count = 0
            for root, dirs, files in os.walk(extract_dir):
                for txt_file in files:
                    if txt_file.endswith('.txt'):
                        seq_name = txt_file[:-4]
                        try:
                            with open(os.path.join(root, txt_file), 'r') as f:
                                content = f.read().strip()
                                att_vec = [int(x.strip()) for x in content.split(',') if x.strip()]
                                DATA_STORE['attributes'][dataset_name][seq_name] = np.array(att_vec)
                                count += 1
                        except:
                            pass
            
            loaded_info.append(f'{dataset_name}: {count}个序列, {len(attr_names)}个属性 ({", ".join(attr_names)})')
        
        if not loaded_info:
            return jsonify({'error': '未找到有效的属性文件，请检查压缩包结构'}), 400
        
        # 验证序列与序列列表的一致性
        verification_results = {}
        all_verified = True
        
        if mode == 'batch':
            for ds_name in DATA_STORE['attributes'].keys():
                if ds_name in DATA_STORE['sequence_lists']:
                    actual_seqs = list(DATA_STORE['attributes'][ds_name].keys())
                    verify_result = verify_sequences_against_list(ds_name, actual_seqs, '属性')
                    verification_results[ds_name] = verify_result
                    if not verify_result.get('verified', True):
                        all_verified = False
        else:
            if dataset_name in DATA_STORE['sequence_lists']:
                actual_seqs = list(DATA_STORE['attributes'][dataset_name].keys())
                verify_result = verify_sequences_against_list(dataset_name, actual_seqs, '属性')
                verification_results[dataset_name] = verify_result
                if not verify_result.get('verified', True):
                    all_verified = False
        
        # 保存数据到磁盘
        save_data_to_disk()
        
        response = {
            'success': True,
            'message': '属性配置上传成功（已自动保存）',
            'details': loaded_info
        }
        
        if verification_results:
            response['verification'] = verification_results
            response['all_verified'] = all_verified
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({'error': f'上传失败: {str(e)}', 'traceback': traceback.format_exc()}), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@app.route('/api/evaluate', methods=['POST'])
def evaluate():
    """运行评测"""
    data = request.get_json()
    
    dataset_name = data.get('dataset_name')
    tracker_names = data.get('tracker_names', [])
    method = data.get('method', 'otb')
    
    if not dataset_name:
        return jsonify({'error': '请选择数据集'}), 400
    
    if dataset_name not in DATA_STORE['ground_truths']:
        return jsonify({'error': f'未找到数据集 {dataset_name} 的Ground Truth'}), 400
    
    if dataset_name not in DATA_STORE['tracker_results']:
        return jsonify({'error': f'未找到数据集 {dataset_name} 的跟踪结果'}), 400
    
    ground_truths = DATA_STORE['ground_truths'][dataset_name]
    
    # 评测前验证数据完整性
    verification_warnings = []
    
    if dataset_name in DATA_STORE['sequence_lists']:
        expected_seqs = DATA_STORE['sequence_lists'][dataset_name]
        
        # 验证GT
        gt_seqs = list(ground_truths.keys())
        gt_verify = verify_sequences_against_list(dataset_name, gt_seqs, 'Ground Truth')
        if not gt_verify.get('verified', True):
            verification_warnings.append({
                'type': 'ground_truth',
                'message': gt_verify['message'],
                'missing': gt_verify.get('missing', []),
                'extra': gt_verify.get('extra', [])
            })
        
        # 验证每个跟踪器的结果
        for trk_name in DATA_STORE['tracker_results'][dataset_name].keys():
            trk_seqs = list(DATA_STORE['tracker_results'][dataset_name][trk_name].keys())
            trk_verify = verify_sequences_against_list(
                dataset_name, trk_seqs, f'跟踪结果({trk_name})')
            if not trk_verify.get('verified', True):
                verification_warnings.append({
                    'type': 'tracker_result',
                    'tracker': trk_name,
                    'message': trk_verify['message'],
                    'missing': trk_verify.get('missing', []),
                    'extra': trk_verify.get('extra', [])
                })
        
        # 验证属性
        if dataset_name in DATA_STORE['attributes']:
            attr_seqs = list(DATA_STORE['attributes'][dataset_name].keys())
            attr_verify = verify_sequences_against_list(dataset_name, attr_seqs, '属性')
            if not attr_verify.get('verified', True):
                verification_warnings.append({
                    'type': 'attributes',
                    'message': attr_verify['message'],
                    'missing': attr_verify.get('missing', []),
                    'extra': attr_verify.get('extra', [])
                })
    
    # 获取跟踪结果
    all_tracker_results = {}
    available_trackers = list(DATA_STORE['tracker_results'][dataset_name].keys())
    
    if not tracker_names:
        tracker_names = available_trackers
    
    for trk_name in tracker_names:
        if trk_name in DATA_STORE['tracker_results'][dataset_name]:
            all_tracker_results[trk_name] = DATA_STORE['tracker_results'][dataset_name][trk_name]
    
    if not all_tracker_results:
        return jsonify({'error': '未找到有效的跟踪结果'}), 400
    
    # 获取属性
    attributes = DATA_STORE['attributes'].get(dataset_name, None)
    attribute_names = DATA_STORE['attribute_names'].get(dataset_name, None)
    
    # 获取时间数据
    tracker_times = DATA_STORE['tracker_times'].get(dataset_name, {})
    
    try:
        results = evaluate_multiple_trackers(
            all_tracker_results,
            ground_truths,
            method=method,
            attributes=attributes,
            attribute_names=attribute_names,
            dataset_name=dataset_name
        )
        
        # 计算FPS
        fps_data = {}
        for trk_name in results.keys():
            if trk_name in tracker_times:
                total_time = 0
                total_frames = 0
                for seq_name, times in tracker_times[trk_name].items():
                    if len(times) > 0:
                        total_time += np.sum(times)
                        total_frames += len(times)
                
                if total_time > 0 and total_frames > 0:
                    fps_data[trk_name] = total_frames / total_time
        
        # 生成session ID并存储结果
        session_id = str(uuid.uuid4())
        DATA_STORE['evaluation_results'][session_id] = {
            'dataset_name': dataset_name,
            'method': method,
            'results': results,
            'attribute_names': attribute_names if attribute_names else [],
            'fps_data': fps_data
        }
        
        # 准备摘要
        summary = {
            'session_id': session_id,
            'dataset_name': dataset_name,
            'method': method,
            'num_sequences': len(ground_truths),
            'trackers': {},
            'has_attributes': attribute_names is not None and len(attribute_names) > 0,
            'attribute_names': attribute_names if attribute_names else [],
            'attribute_results': {}  # 属性评测结果
        }
        
        for trk_name, trk_result in results.items():
            summary['trackers'][trk_name] = {
                'auc': round(trk_result.get('auc', 0), 3 if method == 'otb' else 1),
                'precision_20': round(trk_result.get('precision_20', 0), 3 if method == 'otb' else 1),
                'fps': round(fps_data.get(trk_name, 0), 1) if trk_name in fps_data else None
            }
            if method == 'ortrack':
                summary['trackers'][trk_name]['norm_precision_20'] = round(trk_result.get('norm_precision_20', 0), 1)
                summary['trackers'][trk_name]['success_50'] = round(trk_result.get('success_50', 0), 1)
                summary['trackers'][trk_name]['success_75'] = round(trk_result.get('success_75', 0), 1)
            
            # 提取属性结果
            if 'attribute_results' in trk_result:
                summary['trackers'][trk_name]['has_attribute_results'] = True
                summary['trackers'][trk_name]['attribute_count'] = len(trk_result['attribute_results'])
                
                # 收集每个属性的详细结果
                for attr_name, attr_data in trk_result['attribute_results'].items():
                    if attr_name not in summary['attribute_results']:
                        summary['attribute_results'][attr_name] = {'num_seqs': attr_data.get('num_seqs', 0), 'trackers': {}}
                    
                    summary['attribute_results'][attr_name]['trackers'][trk_name] = {
                        'auc': round(attr_data.get('auc', 0), 3 if method == 'otb' else 1),
                        'precision_20': round(attr_data.get('precision_20', 0), 3 if method == 'otb' else 1)
                    }
            
            # 添加调试信息
            if '_debug_attr_info' in trk_result:
                summary['_debug'] = trk_result['_debug_attr_info']
        
        # 如果有属性配置但没有生成属性结果，添加警告
        if attribute_names and len(attribute_names) > 0:
            first_tracker = list(results.values())[0] if results else {}
            attr_results = first_tracker.get('attribute_results', {})
            if len(attr_results) == 0:
                debug_info = first_tracker.get('_debug_attr_info', {})
                warning_msg = f"警告：检测到属性配置({len(attribute_names)}个属性)，但属性评测结果为空。"
                if debug_info:
                    warning_msg += f"\n调试信息: 属性文件包含{debug_info.get('attr_seq_count', 0)}个序列, "
                    warning_msg += f"GT包含{debug_info.get('gt_seq_count', 0)}个序列, "
                    warning_msg += f"匹配到{debug_info.get('matched_count', 0)}个序列。"
                    if debug_info.get('sample_attr_seqs'):
                        warning_msg += f"\n属性文件序列示例: {debug_info.get('sample_attr_seqs')}"
                    if debug_info.get('sample_gt_seqs'):
                        warning_msg += f"\nGT序列示例: {debug_info.get('sample_gt_seqs')}"
                summary['attribute_warning'] = warning_msg
        
        # 构建响应
        response_data = {
            'success': True,
            'message': f'评测完成！共评测 {len(all_tracker_results)} 个算法在 {len(ground_truths)} 个序列上的表现',
            'data': summary
        }
        
        # 添加验证警告
        if verification_warnings:
            response_data['verification_warnings'] = verification_warnings
            response_data['message'] += f'（注意：有 {len(verification_warnings)} 个数据完整性警告）'
        
        # 禁用缓存
        resp = make_response(jsonify(response_data))
        resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
    
    except Exception as e:
        return jsonify({
            'error': f'评测失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/evaluate_batch', methods=['POST'])
def evaluate_batch():
    """批量评测 - 同时评测多个数据集，计算平均值"""
    data = request.get_json()
    
    dataset_names = data.get('dataset_names', [])
    tracker_names = data.get('tracker_names', [])
    method = data.get('method', 'otb')
    
    if not dataset_names:
        return jsonify({'error': '请选择要评测的数据集'}), 400
    
    try:
        all_dataset_results = {}  # {dataset_name: {tracker_name: result}}
        dataset_summaries = {}    # 每个数据集的摘要
        
        # 对每个数据集进行评测
        for dataset_name in dataset_names:
            if dataset_name not in DATA_STORE['ground_truths']:
                continue
            if dataset_name not in DATA_STORE['tracker_results']:
                continue
            
            ground_truths = DATA_STORE['ground_truths'][dataset_name]
            
            # 获取该数据集的跟踪结果
            ds_tracker_results = {}
            available_trackers = list(DATA_STORE['tracker_results'][dataset_name].keys())
            
            selected_trackers = tracker_names if tracker_names else available_trackers
            
            for trk_name in selected_trackers:
                if trk_name in DATA_STORE['tracker_results'][dataset_name]:
                    ds_tracker_results[trk_name] = DATA_STORE['tracker_results'][dataset_name][trk_name]
            
            if not ds_tracker_results:
                continue
            
            # 获取属性
            attributes = DATA_STORE['attributes'].get(dataset_name, None)
            attribute_names = DATA_STORE['attribute_names'].get(dataset_name, None)
            
            # 评测
            results = evaluate_multiple_trackers(
                ds_tracker_results,
                ground_truths,
                method=method,
                attributes=attributes,
                attribute_names=attribute_names,
                dataset_name=dataset_name
            )
            
            all_dataset_results[dataset_name] = results
            
            # 生成该数据集的摘要
            ds_summary = {
                'num_sequences': len(ground_truths),
                'num_trackers': len(results),
                'trackers': {}
            }
            
            for trk_name, trk_result in results.items():
                ds_summary['trackers'][trk_name] = {
                    'auc': round(trk_result.get('auc', 0), 3 if method == 'otb' else 1),
                    'precision_20': round(trk_result.get('precision_20', 0), 3 if method == 'otb' else 1),
                }
                if method == 'ortrack':
                    ds_summary['trackers'][trk_name]['norm_precision_20'] = round(trk_result.get('norm_precision_20', 0), 1)
            
            dataset_summaries[dataset_name] = ds_summary
        
        if not all_dataset_results:
            return jsonify({'error': '没有有效的评测数据'}), 400
        
        # 计算每个算法在所有数据集上的平均值
        all_trackers = set()
        for ds_results in all_dataset_results.values():
            all_trackers.update(ds_results.keys())
        
        tracker_averages = {}
        for trk_name in all_trackers:
            auc_values = []
            prec_values = []
            norm_prec_values = []
            
            for ds_name, ds_results in all_dataset_results.items():
                if trk_name in ds_results:
                    trk_result = ds_results[trk_name]
                    auc_values.append(trk_result.get('auc', 0))
                    prec_values.append(trk_result.get('precision_20', 0))
                    if method == 'ortrack':
                        norm_prec_values.append(trk_result.get('norm_precision_20', 0))
            
            if auc_values:
                tracker_averages[trk_name] = {
                    'datasets_count': len(auc_values),
                    'avg_auc': round(sum(auc_values) / len(auc_values), 3 if method == 'otb' else 1),
                    'avg_precision_20': round(sum(prec_values) / len(prec_values), 3 if method == 'otb' else 1),
                }
                if method == 'ortrack' and norm_prec_values:
                    tracker_averages[trk_name]['avg_norm_precision_20'] = round(sum(norm_prec_values) / len(norm_prec_values), 1)
        
        # 生成session ID并存储
        session_id = str(uuid.uuid4())
        DATA_STORE['evaluation_results'][session_id] = {
            'type': 'batch',
            'method': method,
            'dataset_names': list(all_dataset_results.keys()),
            'all_results': all_dataset_results,
            'averages': tracker_averages
        }
        
        return jsonify({
            'success': True,
            'message': f'批量评测完成！评测了 {len(all_dataset_results)} 个数据集，{len(all_trackers)} 个算法',
            'data': {
                'session_id': session_id,
                'method': method,
                'datasets': dataset_summaries,
                'averages': tracker_averages
            }
        })
    
    except Exception as e:
        return jsonify({
            'error': f'批量评测失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/export_batch_excel', methods=['POST'])
def export_batch_excel():
    """导出批量评测结果为Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel导出功能不可用，请安装openpyxl'}), 500
    
    data = request.get_json()
    session_id = data.get('session_id')
    
    if not session_id or session_id not in DATA_STORE['evaluation_results']:
        return jsonify({'error': '评测结果不存在'}), 400
    
    session_data = DATA_STORE['evaluation_results'][session_id]
    
    if session_data.get('type') != 'batch':
        return jsonify({'error': '该结果不是批量评测结果'}), 400
    
    method = session_data['method']
    dataset_names = session_data['dataset_names']
    all_results = session_data['all_results']
    averages = session_data['averages']
    
    temp_dir = tempfile.mkdtemp()
    try:
        wb = Workbook()
        
        # 样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        avg_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ========== 平均值汇总表 ==========
        ws_avg = wb.active
        ws_avg.title = '平均值汇总'
        
        # 表头
        if method == 'otb':
            headers = ['算法', '数据集数', '平均AUC', '平均Precision@20']
        else:
            headers = ['算法', '数据集数', '平均AUC', '平均Precision@20', '平均Norm Prec@20']
        
        for col, header in enumerate(headers, 1):
            cell = ws_avg.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 按平均AUC排序
        sorted_trackers = sorted(averages.items(), key=lambda x: x[1]['avg_auc'], reverse=True)
        
        for row, (trk_name, trk_avg) in enumerate(sorted_trackers, 2):
            if method == 'otb':
                values = [trk_name, trk_avg['datasets_count'], trk_avg['avg_auc'], trk_avg['avg_precision_20']]
            else:
                values = [trk_name, trk_avg['datasets_count'], trk_avg['avg_auc'], 
                          trk_avg['avg_precision_20'], trk_avg.get('avg_norm_precision_20', '-')]
            
            for col, value in enumerate(values, 1):
                cell = ws_avg.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        for col in range(1, len(headers) + 1):
            ws_avg.column_dimensions[get_column_letter(col)].width = 18
        
        # ========== 每个数据集的详细结果 ==========
        for ds_name in dataset_names:
            ds_results = all_results.get(ds_name, {})
            if not ds_results:
                continue
            
            # 创建工作表（名称最多31个字符）
            sheet_name = ds_name[:31] if len(ds_name) > 31 else ds_name
            ws_ds = wb.create_sheet(sheet_name)
            
            # 表头
            if method == 'otb':
                headers = ['算法', 'AUC', 'Precision@20']
            else:
                headers = ['算法', 'AUC', 'Precision@20', 'Norm Prec@20']
            
            for col, header in enumerate(headers, 1):
                cell = ws_ds.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            sorted_ds_trackers = sorted(ds_results.items(), key=lambda x: x[1].get('auc', 0), reverse=True)
            
            for row, (trk_name, trk_result) in enumerate(sorted_ds_trackers, 2):
                auc = trk_result.get('auc', 0)
                prec = trk_result.get('precision_20', 0)
                
                if method == 'otb':
                    values = [trk_name, round(auc, 3), round(prec, 3)]
                else:
                    norm_prec = trk_result.get('norm_precision_20', 0)
                    values = [trk_name, round(auc, 1), round(prec, 1), round(norm_prec, 1)]
                
                for col, value in enumerate(values, 1):
                    cell = ws_ds.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for col in range(1, len(headers) + 1):
                ws_ds.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== 综合对比表（算法×数据集）==========
        ws_compare = wb.create_sheet('综合对比(AUC)')
        
        # 表头：算法 + 各数据集 + 平均
        compare_headers = ['算法'] + dataset_names + ['平均']
        for col, header in enumerate(compare_headers, 1):
            cell = ws_compare.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = avg_fill if col == len(compare_headers) else header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row, (trk_name, trk_avg) in enumerate(sorted_trackers, 2):
            # 算法名
            cell = ws_compare.cell(row=row, column=1, value=trk_name)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # 每个数据集的AUC
            for col, ds_name in enumerate(dataset_names, 2):
                ds_results = all_results.get(ds_name, {})
                if trk_name in ds_results:
                    auc = ds_results[trk_name].get('auc', 0)
                    value = round(auc, 3 if method == 'otb' else 1)
                else:
                    value = '-'
                cell = ws_compare.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
            
            # 平均值
            cell = ws_compare.cell(row=row, column=len(compare_headers), value=trk_avg['avg_auc'])
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        for col in range(1, len(compare_headers) + 1):
            ws_compare.column_dimensions[get_column_letter(col)].width = 12
        
        # 保存
        excel_path = os.path.join(temp_dir, 'batch_results.xlsx')
        wb.save(excel_path)
        
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='batch_benchmark_results.xlsx'
        )
    
    except Exception as e:
        return jsonify({
            'error': f'导出Excel失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500
    finally:
        import threading
        def cleanup():
            import time
            time.sleep(10)
            shutil.rmtree(temp_dir, ignore_errors=True)
        threading.Thread(target=cleanup).start()


@app.route('/api/plot', methods=['POST'])
def generate_plot():
    """生成图表"""
    data = request.get_json()
    
    session_id = data.get('session_id')
    plot_type = data.get('plot_type', 'success')
    attribute_name = data.get('attribute_name', None)
    rank_num = data.get('rank_num', 10)
    legend_fontsize = data.get('legend_fontsize', 10)
    
    # 自定义颜色
    tracker_colors = data.get('tracker_colors', None)
    
    # 雷达图参数
    radar_attributes = data.get('radar_attributes', [])
    radar_metric = data.get('radar_metric', 'auc')
    
    # IoU曲线参数
    seq_name = data.get('seq_name', None)
    selected_trackers = data.get('selected_trackers', [])
    
    if not session_id:
        return jsonify({'error': '缺少session_id参数'}), 400
    
    if session_id not in DATA_STORE['evaluation_results']:
        return jsonify({'error': '评测结果不存在，请先进行评测'}), 400
    
    session_data = DATA_STORE['evaluation_results'][session_id]
    tracker_results = session_data['results']
    dataset_name = session_data['dataset_name']
    method = session_data['method']
    stored_attribute_names = session_data.get('attribute_names', [])
    
    try:
        img = None
        plot_info = ""
        
        if plot_type == 'success':
            if attribute_name:
                # 属性成功率图
                img = plot_attribute_success(
                    tracker_results, attribute_name,
                    rank_num=rank_num, legend_fontsize=legend_fontsize,
                    tracker_colors=tracker_colors
                )
                plot_info = f"属性 '{attribute_name}' 的成功率曲线"
            else:
                # 总体成功率图
                img = plot_success_curve(
                    tracker_results,
                    title=f'Success plots of OPE - {dataset_name}',
                    rank_num=rank_num, legend_fontsize=legend_fontsize,
                    tracker_colors=tracker_colors
                )
                plot_info = "总体成功率曲线"
        
        elif plot_type == 'precision':
            if attribute_name:
                # 属性精度图
                img = plot_attribute_precision(
                    tracker_results, attribute_name,
                    rank_num=rank_num, legend_fontsize=legend_fontsize,
                    tracker_colors=tracker_colors
                )
                plot_info = f"属性 '{attribute_name}' 的精度曲线"
            else:
                # 总体精度图
                img = plot_precision_curve(
                    tracker_results,
                    title=f'Precision plots of OPE - {dataset_name}',
                    rank_num=rank_num, legend_fontsize=legend_fontsize,
                    tracker_colors=tracker_colors
                )
                plot_info = "总体精度曲线"
        
        elif plot_type == 'norm_precision':
            if method != 'ortrack':
                return jsonify({'error': '归一化精度曲线仅适用于ORTrack方法'}), 400
            
            img = plot_norm_precision_curve(
                tracker_results,
                title=f'Normalized Precision plots - {dataset_name}',
                rank_num=rank_num, legend_fontsize=legend_fontsize,
                tracker_colors=tracker_colors
            )
            plot_info = "归一化精度曲线"
        
        elif plot_type == 'radar':
            if not radar_attributes:
                radar_attributes = stored_attribute_names
            
            if not radar_attributes or len(radar_attributes) < 3:
                return jsonify({'error': '雷达图至少需要选择3个属性'}), 400
            
            img = plot_radar_chart(
                tracker_results,
                radar_attributes,
                metric=radar_metric,
                title=f'Radar Chart - {dataset_name}',
                tracker_colors=tracker_colors,
                legend_fontsize=legend_fontsize
            )
            plot_info = f"雷达图 (指标: {radar_metric}, 属性: {', '.join(radar_attributes)})"
        
        elif plot_type == 'iou_curve':
            # IoU曲线对比图
            if not seq_name:
                return jsonify({'error': '请选择要对比的视频序列'}), 400
            
            if not selected_trackers:
                return jsonify({'error': '请选择要对比的算法'}), 400
            
            # 获取GT和跟踪结果
            if dataset_name not in DATA_STORE['ground_truths']:
                return jsonify({'error': f'未找到数据集 {dataset_name} 的Ground Truth'}), 400
            
            gt_data = DATA_STORE['ground_truths'][dataset_name]
            if seq_name not in gt_data:
                return jsonify({'error': f'未找到序列 {seq_name} 的Ground Truth'}), 400
            
            gt_bb = gt_data[seq_name]
            
            # 计算每个选中算法的IoU
            iou_data = {}
            for trk_name in selected_trackers:
                if dataset_name in DATA_STORE['tracker_results']:
                    if trk_name in DATA_STORE['tracker_results'][dataset_name]:
                        if seq_name in DATA_STORE['tracker_results'][dataset_name][trk_name]:
                            pred_bb = DATA_STORE['tracker_results'][dataset_name][trk_name][seq_name]
                            iou_per_frame = compute_sequence_iou(pred_bb, gt_bb)
                            iou_data[trk_name] = iou_per_frame
            
            if not iou_data:
                return jsonify({'error': '所选算法在该序列上没有跟踪结果'}), 400
            
            img = plot_sequence_iou_curve(
                iou_data, seq_name,
                title=f'IoU Curve - {seq_name}',
                tracker_colors=tracker_colors,
                legend_fontsize=legend_fontsize
            )
            plot_info = f"序列 '{seq_name}' 的IoU曲线对比 ({len(iou_data)}个算法)"
        
        elif plot_type == 'center_error_curve':
            # 中心误差曲线对比图
            if not seq_name:
                return jsonify({'error': '请选择要对比的视频序列'}), 400
            
            if not selected_trackers:
                return jsonify({'error': '请选择要对比的算法'}), 400
            
            if dataset_name not in DATA_STORE['ground_truths']:
                return jsonify({'error': f'未找到数据集 {dataset_name} 的Ground Truth'}), 400
            
            gt_data = DATA_STORE['ground_truths'][dataset_name]
            if seq_name not in gt_data:
                return jsonify({'error': f'未找到序列 {seq_name} 的Ground Truth'}), 400
            
            gt_bb = gt_data[seq_name]
            
            error_data = {}
            for trk_name in selected_trackers:
                if dataset_name in DATA_STORE['tracker_results']:
                    if trk_name in DATA_STORE['tracker_results'][dataset_name]:
                        if seq_name in DATA_STORE['tracker_results'][dataset_name][trk_name]:
                            pred_bb = DATA_STORE['tracker_results'][dataset_name][trk_name][seq_name]
                            error_per_frame = compute_sequence_center_error(pred_bb, gt_bb)
                            error_data[trk_name] = error_per_frame
            
            if not error_data:
                return jsonify({'error': '所选算法在该序列上没有跟踪结果'}), 400
            
            img = plot_sequence_center_error_curve(
                error_data, seq_name,
                title=f'Center Error Curve - {seq_name}',
                tracker_colors=tracker_colors,
                legend_fontsize=legend_fontsize
            )
            plot_info = f"序列 '{seq_name}' 的中心误差曲线对比 ({len(error_data)}个算法)"
        
        else:
            return jsonify({'error': f'未知的图表类型: {plot_type}'}), 400
        
        if img is None:
            # 详细的错误信息
            if plot_type in ['success', 'precision'] and attribute_name:
                # 检查属性是否存在
                has_attr = False
                for trk_name, trk_result in tracker_results.items():
                    if 'attribute_results' in trk_result and attribute_name in trk_result['attribute_results']:
                        has_attr = True
                        break
                
                if not has_attr:
                    available_attrs = set()
                    for trk_result in tracker_results.values():
                        if 'attribute_results' in trk_result:
                            available_attrs.update(trk_result['attribute_results'].keys())
                    
                    if available_attrs:
                        return jsonify({
                            'error': f"未找到属性 '{attribute_name}' 的评测结果。可用属性: {', '.join(sorted(available_attrs))}"
                        }), 400
                    else:
                        return jsonify({
                            'error': '评测结果中没有属性数据，请确保已上传属性配置并重新进行评测'
                        }), 400
            
            elif plot_type == 'radar':
                # 检查雷达图数据
                missing_attrs = []
                for attr in radar_attributes:
                    found = False
                    for trk_result in tracker_results.values():
                        if 'attribute_results' in trk_result and attr in trk_result['attribute_results']:
                            found = True
                            break
                    if not found:
                        missing_attrs.append(attr)
                
                if missing_attrs:
                    return jsonify({
                        'error': f"以下属性没有评测数据: {', '.join(missing_attrs)}。请检查属性配置是否正确上传"
                    }), 400
            
            return jsonify({'error': '生成图表失败，请检查数据是否完整'}), 400
        
        return jsonify({
            'success': True,
            'message': f'{plot_info} 生成成功',
            'image': img
        })
    
    except Exception as e:
        return jsonify({
            'error': f'生成图表失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/export', methods=['POST'])
def export_results():
    """导出所有结果"""
    data = request.get_json()
    session_id = data.get('session_id')
    rank_num = data.get('rank_num', 10)
    legend_fontsize = data.get('legend_fontsize', 10)
    
    if not session_id or session_id not in DATA_STORE['evaluation_results']:
        return jsonify({'error': '评测结果不存在'}), 400
    
    session_data = DATA_STORE['evaluation_results'][session_id]
    tracker_results = session_data['results']
    dataset_name = session_data['dataset_name']
    method = session_data['method']
    attribute_names = session_data.get('attribute_names', [])
    
    temp_dir = tempfile.mkdtemp()
    try:
        # 生成所有图表
        plots = generate_all_plots(
            tracker_results,
            attribute_names=attribute_names,
            output_dir=temp_dir,
            dataset_name=dataset_name,
            method=method,
            legend_fontsize=legend_fontsize,
            rank_num=rank_num
        )
        
        # 保存JSON结果
        json_path = os.path.join(temp_dir, 'results.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'dataset_name': dataset_name,
                'method': method,
                'attribute_names': attribute_names,
                'results': tracker_results
            }, f, ensure_ascii=False, indent=2, default=str)
        
        # 创建ZIP
        zip_path = os.path.join(temp_dir, 'export.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(json_path, 'results.json')
            for name, path in plots.items():
                if os.path.exists(path):
                    zf.write(path, f'plots/{os.path.basename(path)}')
        
        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'benchmark_results_{dataset_name}.zip'
        )
    
    except Exception as e:
        return jsonify({
            'error': f'导出失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500
    finally:
        # 延迟清理
        import threading
        def cleanup():
            import time
            time.sleep(10)
            shutil.rmtree(temp_dir, ignore_errors=True)
        threading.Thread(target=cleanup).start()


@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    """导出评测结果为Excel文件"""
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Excel导出功能不可用，请安装openpyxl: pip install openpyxl'}), 500
    
    data = request.get_json()
    session_id = data.get('session_id')
    export_type = data.get('export_type', 'all')  # 'all', 'overall', 'attribute'
    
    if not session_id or session_id not in DATA_STORE['evaluation_results']:
        return jsonify({'error': '评测结果不存在'}), 400
    
    session_data = DATA_STORE['evaluation_results'][session_id]
    tracker_results = session_data['results']
    dataset_name = session_data['dataset_name']
    method = session_data['method']
    attribute_names = session_data.get('attribute_names', [])
    fps_data = session_data.get('fps_data', {})
    
    temp_dir = tempfile.mkdtemp()
    try:
        wb = Workbook()
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== 总体结果表 ==========
        ws_overall = wb.active
        ws_overall.title = '总体结果'
        
        # 表头
        if method == 'otb':
            headers = ['算法', 'AUC', 'Precision@20', 'FPS']
        else:
            headers = ['算法', 'AUC', 'Precision@20', 'Norm Prec@20', 'OP50', 'OP75', 'FPS']
        
        for col, header in enumerate(headers, 1):
            cell = ws_overall.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据行（按AUC排序）
        sorted_trackers = sorted(tracker_results.items(), 
                                  key=lambda x: x[1].get('auc', 0), reverse=True)
        
        for row, (trk_name, trk_result) in enumerate(sorted_trackers, 2):
            auc = trk_result.get('auc', 0)
            prec = trk_result.get('precision_20', 0)
            fps = fps_data.get(trk_name, None)
            
            if method == 'otb':
                values = [trk_name, round(auc, 3), round(prec, 3), round(fps, 1) if fps else '-']
            else:
                norm_prec = trk_result.get('norm_precision_20', 0)
                op50 = trk_result.get('success_50', 0)
                op75 = trk_result.get('success_75', 0)
                values = [trk_name, round(auc, 1), round(prec, 1), round(norm_prec, 1), 
                          round(op50, 1), round(op75, 1), round(fps, 1) if fps else '-']
            
            for col, value in enumerate(values, 1):
                cell = ws_overall.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws_overall.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== 属性结果表 ==========
        if attribute_names:
            # 属性AUC表
            ws_attr_auc = wb.create_sheet('属性AUC')
            
            # 表头：算法名 + 各属性
            attr_headers = ['算法'] + attribute_names
            for col, header in enumerate(attr_headers, 1):
                cell = ws_attr_auc.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row, (trk_name, trk_result) in enumerate(sorted_trackers, 2):
                cell = ws_attr_auc.cell(row=row, column=1, value=trk_name)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                attr_results = trk_result.get('attribute_results', {})
                for col, attr_name in enumerate(attribute_names, 2):
                    if attr_name in attr_results:
                        value = round(attr_results[attr_name].get('auc', 0), 3 if method == 'otb' else 1)
                    else:
                        value = '-'
                    cell = ws_attr_auc.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for col in range(1, len(attr_headers) + 1):
                ws_attr_auc.column_dimensions[get_column_letter(col)].width = 12
            
            # 属性Precision表
            ws_attr_prec = wb.create_sheet('属性Precision')
            
            for col, header in enumerate(attr_headers, 1):
                cell = ws_attr_prec.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            for row, (trk_name, trk_result) in enumerate(sorted_trackers, 2):
                cell = ws_attr_prec.cell(row=row, column=1, value=trk_name)
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                attr_results = trk_result.get('attribute_results', {})
                for col, attr_name in enumerate(attribute_names, 2):
                    if attr_name in attr_results:
                        value = round(attr_results[attr_name].get('precision_20', 0), 3 if method == 'otb' else 1)
                    else:
                        value = '-'
                    cell = ws_attr_prec.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border
            
            for col in range(1, len(attr_headers) + 1):
                ws_attr_prec.column_dimensions[get_column_letter(col)].width = 12
        
        # 保存Excel
        excel_path = os.path.join(temp_dir, f'results_{dataset_name}.xlsx')
        wb.save(excel_path)
        
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'benchmark_results_{dataset_name}.xlsx'
        )
    
    except Exception as e:
        return jsonify({
            'error': f'导出Excel失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500
    finally:
        import threading
        def cleanup():
            import time
            time.sleep(10)
            shutil.rmtree(temp_dir, ignore_errors=True)
        threading.Thread(target=cleanup).start()


@app.route('/api/draw_bbox', methods=['POST'])
def draw_bbox():
    """绘制跟踪框到图像上"""
    try:
        if not CV2_AVAILABLE:
            return jsonify({'error': '跟踪框绘制功能不可用，请安装opencv-python: pip install opencv-python-headless'}), 500
        
        import cv2
        
        data = request.get_json()
        if not data:
            return jsonify({'error': '请求数据为空'}), 400
            
        dataset_name = data.get('dataset_name')
        tracker_names = data.get('tracker_names', [])
        seq_names = data.get('seq_names', [])
        image_base_path = data.get('image_base_path', '').strip()
        output_base_path = data.get('output_base_path', '').strip()
        line_width = data.get('line_width', 2)
        draw_gt = data.get('draw_gt', True)
        custom_colors = data.get('tracker_colors') or {}
        
        # 参数验证
        if not dataset_name:
            return jsonify({'error': '请选择数据集'}), 400
        if not tracker_names:
            return jsonify({'error': '请选择要绘制的算法'}), 400
        if not seq_names:
            return jsonify({'error': '请选择要绘制的序列'}), 400
        if not image_base_path:
            return jsonify({'error': '请输入图像根目录路径'}), 400
        if not output_base_path:
            return jsonify({'error': '请输入输出目录路径'}), 400
        
        # 检查图像路径
        if not os.path.exists(image_base_path):
            return jsonify({'error': f'图像根目录不存在: {image_base_path}'}), 400
        
        if not os.path.isdir(image_base_path):
            return jsonify({'error': f'图像路径不是目录: {image_base_path}'}), 400
        
        # 创建输出目录
        try:
            os.makedirs(output_base_path, exist_ok=True)
        except Exception as e:
            return jsonify({'error': f'无法创建输出目录: {output_base_path}, 原因: {str(e)}'}), 400
        
        # 预定义颜色 (BGR格式)
        colors = [
            (0, 255, 0),    # 绿色 - GT
            (0, 0, 255),    # 红色
            (255, 0, 0),    # 蓝色
            (0, 255, 255),  # 黄色
            (255, 0, 255),  # 紫色
            (255, 255, 0),  # 青色
            (0, 128, 255),  # 橙色
            (128, 0, 255),  # 粉色
            (255, 128, 0),  # 天蓝
            (128, 255, 0),  # 黄绿
        ]
        
        results = []
        errors = []
        
        for seq_name in seq_names:
            try:
                # 查找图像目录 - 尝试多种可能的路径
                seq_image_path = None
                possible_paths = [
                    os.path.join(image_base_path, seq_name),
                    os.path.join(image_base_path, seq_name, 'img'),
                    os.path.join(image_base_path, seq_name, 'imgs'),
                    os.path.join(image_base_path, seq_name, 'frames'),
                ]
                
                for path in possible_paths:
                    if os.path.exists(path) and os.path.isdir(path):
                        # 检查是否有图像文件
                        files = os.listdir(path)
                        if any(f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')) for f in files):
                            seq_image_path = path
                            break
                
                if seq_image_path is None:
                    errors.append(f'序列 {seq_name}: 未找到图像目录，尝试了: {", ".join(possible_paths[:2])}')
                    continue
                
                # 获取图像文件列表
                image_files = sorted([f for f in os.listdir(seq_image_path) 
                                      if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp'))])
                
                if not image_files:
                    errors.append(f'序列 {seq_name}: 目录存在但没有图像文件')
                    continue
                
                # 获取GT
                gt_bb = None
                if draw_gt and dataset_name in DATA_STORE['ground_truths']:
                    gt_bb = DATA_STORE['ground_truths'][dataset_name].get(seq_name)
                
                # 获取各算法的跟踪结果
                tracker_bbs = {}
                for trk_name in tracker_names:
                    if dataset_name in DATA_STORE['tracker_results']:
                        if trk_name in DATA_STORE['tracker_results'][dataset_name]:
                            bb = DATA_STORE['tracker_results'][dataset_name][trk_name].get(seq_name)
                            if bb is not None:
                                tracker_bbs[trk_name] = bb
                
                if not tracker_bbs and gt_bb is None:
                    errors.append(f'序列 {seq_name}: 没有找到GT或跟踪结果')
                    continue
                
                # 创建输出目录
                seq_output_path = os.path.join(output_base_path, seq_name)
                os.makedirs(seq_output_path, exist_ok=True)
                
                # 绘制每一帧
                num_frames = len(image_files)
                drawn_frames = 0
                
                for frame_idx, img_file in enumerate(image_files):
                    img_path = os.path.join(seq_image_path, img_file)
                    img = cv2.imread(img_path)
                    
                    if img is None:
                        continue
                    
                    # 绘制GT (绿色) - 只绘制框，不显示文字
                    if gt_bb is not None and frame_idx < len(gt_bb):
                        box = gt_bb[frame_idx]
                        if not np.isnan(box).any() and box[2] > 0 and box[3] > 0:
                            x, y, w, h = int(box[0]), int(box[1]), int(box[2]), int(box[3])
                            cv2.rectangle(img, (x, y), (x + w, y + h), colors[0], line_width)
                    
                    # 绘制各算法的跟踪框 - 只绘制框，不显示跟踪器名字
                    for i, (trk_name, trk_bb) in enumerate(tracker_bbs.items()):
                        if frame_idx < len(trk_bb):
                            box = trk_bb[frame_idx]
                            if not np.isnan(box).any() and box[2] > 0 and box[3] > 0:
                                # 获取颜色
                                if trk_name in custom_colors and custom_colors[trk_name]:
                                    hex_color = custom_colors[trk_name].lstrip('#')
                                    try:
                                        r = int(hex_color[0:2], 16)
                                        g = int(hex_color[2:4], 16)
                                        b = int(hex_color[4:6], 16)
                                        color = (b, g, r)  # BGR
                                    except:
                                        color = colors[(i + 1) % len(colors)]
                                else:
                                    color = colors[(i + 1) % len(colors)]
                                
                                x, y, w, h = int(box[0]), int(box[1]), int(box[2]), int(box[3])
                                cv2.rectangle(img, (x, y), (x + w, y + h), color, line_width)
                    
                    # 保存图像
                    output_file = os.path.join(seq_output_path, img_file)
                    cv2.imwrite(output_file, img)
                    drawn_frames += 1
                
                results.append({
                    'seq_name': seq_name,
                    'num_frames': drawn_frames,
                    'output_path': seq_output_path,
                    'trackers': list(tracker_bbs.keys())
                })
                
            except Exception as e:
                errors.append(f'序列 {seq_name}: 处理出错 - {str(e)}')
        
        if not results:
            return jsonify({
                'error': '没有成功绘制任何序列',
                'details': errors
            }), 400
        
        return jsonify({
            'success': True,
            'message': f'成功绘制 {len(results)} 个序列',
            'results': results,
            'errors': errors if errors else None
        })
        
    except Exception as e:
        return jsonify({
            'error': f'绘制跟踪框失败: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500


@app.route('/api/get_sequences_with_images', methods=['POST'])
def get_sequences_with_images():
    """获取指定路径下有图像的序列列表"""
    data = request.get_json()
    image_base_path = data.get('image_base_path', '')
    dataset_name = data.get('dataset_name', '')
    
    if not image_base_path or not os.path.exists(image_base_path):
        return jsonify({'error': '图像目录不存在'}), 400
    
    # 获取数据集中的序列
    if dataset_name and dataset_name in DATA_STORE['ground_truths']:
        dataset_seqs = set(DATA_STORE['ground_truths'][dataset_name].keys())
    else:
        dataset_seqs = set()
    
    # 检查哪些序列有图像
    available_seqs = []
    for item in os.listdir(image_base_path):
        item_path = os.path.join(image_base_path, item)
        if os.path.isdir(item_path):
            # 检查目录下是否有图像
            has_images = False
            for sub in os.listdir(item_path):
                if sub.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
                    has_images = True
                    break
                # 检查img子目录
                sub_path = os.path.join(item_path, sub)
                if os.path.isdir(sub_path) and sub.lower() == 'img':
                    for f in os.listdir(sub_path):
                        if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp')):
                            has_images = True
                            break
            
            if has_images:
                available_seqs.append({
                    'name': item,
                    'in_dataset': item in dataset_seqs
                })
    
    return jsonify({
        'sequences': sorted(available_seqs, key=lambda x: x['name'])
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
